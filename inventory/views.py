import io
import json
import random
import pytz
import openpyxl
import traceback
import re
from datetime import datetime, date as _date
from django.core.cache import cache
from django.db.models import Sum, F, ExpressionWrapper, IntegerField
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.core.exceptions import PermissionDenied
from django.http import JsonResponse, HttpResponse
from django.utils import timezone
from django.views.decorators.http import require_POST
from django.views.decorators.csrf import ensure_csrf_cookie, csrf_exempt
from django.views.decorators.http import require_http_methods
from django.contrib import messages
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from .models import Item, Transaction, BorrowRequest, DeviceMonitor, TransactionDevice
from .forms import ItemForm, StaffBorrowForm, TransactionConditionForm, BorrowRequestForm
from .decorators import no_cache
from .broadcasts import broadcast_device_monitoring, broadcast_dashboard
from .utils import get_ph_time, format_ph_time, parse_excel_date, PH_TZ   # <-- only place these come from


def _broadcasts():
    from inventory import broadcasts as b
    return b


# ─────────────────────────────────────────────────────────────────────────────
#  Public / unauthenticated
# ─────────────────────────────────────────────────────────────────────────────

def welcome(request):
    if request.user.is_authenticated:
        return redirect('index')

    borrow_form     = BorrowRequestForm()
    borrow_success  = None
    generated_tx_id = str(random.randint(10000, 99999))

    if 'borrow_success' in request.session:
        borrow_success  = request.session.pop('borrow_success')
        generated_tx_id = str(random.randint(10000, 99999))

    if request.method == 'POST' and request.POST.get('action') == 'borrow_request':
        borrow_form = BorrowRequestForm(request.POST)
        if borrow_form.is_valid():
            req   = borrow_form.save(commit=False)
            tx_id = request.POST.get('transaction_id', str(random.randint(10000, 99999)))
            while BorrowRequest.objects.filter(transaction_id=tx_id).exists():
                tx_id = str(random.randint(10000, 99999))
            req.transaction_id = tx_id
            req.save()

            request.session['borrow_success'] = req.transaction_id
            b = _broadcasts()
            b.broadcast_borrow_requests()
            b.broadcast_dashboard()
            return redirect('welcome')

    return render(request, 'inventory/welcome.html', {
        'borrow_form':     borrow_form,
        'borrow_success':  borrow_success,
        'generated_tx_id': generated_tx_id,
        'available_items': Item.objects.filter(available_quantity__gt=0),
    })

# ─────────────────────────────────────────────────────────────────────────────
#  Dashboard
# ─────────────────────────────────────────────────────────────────────────────

@login_required
@no_cache
def index(request):
    from django.db.models import Count, Q, Sum

    pending_count  = BorrowRequest.objects.filter(status='pending').count()
    items          = Item.objects.all()
    available_qty  = Item.objects.aggregate(t=Sum('available_quantity'))['t'] or 0

    agg = Transaction.objects.annotate(
        still_out=ExpressionWrapper(
            F('quantity_borrowed') - F('returned_qty'),
            output_field=IntegerField()
        )
    ).aggregate(total=Sum('still_out'))
    borrowed_qty = max(0, agg['total'] or 0)

    # Stats from DeviceMonitor (instead of Transaction counts)
    dm_counts = DeviceMonitor.objects.aggregate(
        dm_returned=Count('id', filter=Q(date_returned__isnull=False)),
        dm_released=Count('id', filter=Q(is_released=True, date_returned__isnull=True)),
    )
    total_devices  = DeviceMonitor.objects.count()
    active_borrows = dm_counts['dm_released']   # devices currently released
    total_returns  = dm_counts['dm_returned']   # devices physically returned

    # Bar chart data (unchanged)
    offices = list(
        DeviceMonitor.objects.values_list('office_college', flat=True)
        .distinct().order_by('office_college')
    )
    dm_svc = dm_non = dm_seal = dm_miss = dm_inc = []
    if offices:
        agg_by_office = DeviceMonitor.objects.values('office_college').annotate(
            svc=Count('id', filter=Q(serviceable=True)),
            non=Count('id', filter=Q(non_serviceable=True)),
            seal=Count('id', filter=Q(sealed=True)),
            miss=Count('id', filter=Q(missing=True)),
            inc=Count('id', filter=Q(incomplete=True)),
        ).order_by('office_college')
        agg_map = {r['office_college']: r for r in agg_by_office}
        dm_svc  = [agg_map.get(o, {}).get('svc',  0) for o in offices]
        dm_non  = [agg_map.get(o, {}).get('non',  0) for o in offices]
        dm_seal = [agg_map.get(o, {}).get('seal', 0) for o in offices]
        dm_miss = [agg_map.get(o, {}).get('miss', 0) for o in offices]
        dm_inc  = [agg_map.get(o, {}).get('inc',  0) for o in offices]

    return render(request, 'inventory/index.html', {
        'items':          items,
        'total_devices':  total_devices,
        'active_borrows': active_borrows,
        'total_returns':  total_returns,
        'pending_count':  pending_count,
        'available_qty':  available_qty,
        'borrowed_qty':   borrowed_qty,
        'dm_released':    dm_counts['dm_released'],
        'dm_returned':    dm_counts['dm_returned'],
        'dm_offices':     json.dumps(offices),
        'dm_serviceable': json.dumps(dm_svc),
        'dm_non_service': json.dumps(dm_non),
        'dm_sealed':      json.dumps(dm_seal),
        'dm_missing':     json.dumps(dm_miss),
        'dm_incomplete':  json.dumps(dm_inc),
    })

@login_required
def transaction_devices_json(request, transaction_id):
    if request.user.role != 'staff':
        return JsonResponse({'error': 'Forbidden'}, status=403)

    tx = get_object_or_404(Transaction, id=transaction_id)
    devices = list(tx.devices.all())

    if not devices and tx.serial_number:
        serials = [s.strip() for s in tx.serial_number.split(',') if s.strip()]
        dm_map = {}
        for dm in DeviceMonitor.objects.filter(serial_number__in=serials):
            dm_map[dm.serial_number] = dm.box_number

        data = []
        for sn in serials:
            data.append({
                'id':            None,
                'serial_number': sn,
                'box_number':    dm_map.get(sn, '—'),
                'returned':      False,
                'returned_at':   None,
            })
        return JsonResponse({'devices': data})

    data = []
    for d in devices:
        data.append({
            'id':            d.id,
            'serial_number': d.serial_number,
            'box_number':    d.box_number or '—',
            'returned':      d.returned,
            'returned_at':   format_ph_time(d.returned_at),
        })
    return JsonResponse({'devices': data})


# ─────────────────────────────────────────────────────────────────────────────
#  AJAX poll endpoints
# ─────────────────────────────────────────────────────────────────────────────

@login_required
def ajax_dashboard_data(request):
    from django.core.cache import cache
    from inventory.consumers import _build_dashboard_payload
    key = 'dashboard_stats'
    data = cache.get(key)
    if data is None:
        data = _build_dashboard_payload()
        cache.set(key, data, 60)
    return JsonResponse(data)


@login_required
def ajax_borrow_management_data(request):
    if request.user.role != 'staff':
        return JsonResponse({'error': 'Forbidden'}, status=403)
    from django.core.cache import cache
    from inventory.consumers import _build_borrow_management_payload
    key = 'ajax_borrow_mgmt'
    data = cache.get(key)
    if data is None:
        data = _build_borrow_management_payload()
        cache.set(key, data, 30)
    return JsonResponse(data)


@login_required
def ajax_borrow_requests_data(request):
    if request.user.role != 'staff':
        return JsonResponse({'error': 'Forbidden'}, status=403)
    from django.core.cache import cache
    from inventory.consumers import _build_borrow_requests_payload
    key = 'ajax_borrow_requests'
    data = cache.get(key)
    if data is None:
        data = _build_borrow_requests_payload()
        cache.set(key, data, 30)
    return JsonResponse(data)


@login_required
def ajax_device_monitoring_data(request):
    if request.user.role != 'staff':
        return JsonResponse({'error': 'Forbidden'}, status=403)
    from django.core.cache import cache
    from inventory.consumers import _build_device_monitoring_payload
    key = 'ajax_device_monitoring'
    data = cache.get(key)
    if data is None:
        data = _build_device_monitoring_payload()
        cache.set(key, data, 30)
    return JsonResponse(data)


# ─────────────────────────────────────────────────────────────────────────────
#  Item management
# ─────────────────────────────────────────────────────────────────────────────

@login_required
@no_cache
def add_item(request):
    if request.user.role != 'admin':
        raise PermissionDenied
    form = ItemForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        item = form.save(commit=False)
        item.available_quantity = item.quantity
        item.save()
        b = _broadcasts()
        b.broadcast_dashboard()
        b.broadcast_borrow_management()
        return redirect('index')
    return render(request, 'inventory/add_item.html', {'form': form})


@login_required
def edit_item(request, item_id):
    if request.user.role != 'admin':
        messages.error(request, 'You do not have permission to edit items.')
        return redirect('index')

    item = get_object_or_404(Item, id=item_id)

    if request.method == 'POST':
        new_quantity = request.POST.get('available_quantity')
        if new_quantity is not None:
            try:
                new_quantity = int(new_quantity)
                if new_quantity >= 0:
                    item.available_quantity = new_quantity
                    item.save()
                    messages.success(request, f'Updated {item.name} to {item.available_quantity} units.')
                    b = _broadcasts()
                    b.broadcast_dashboard()
                    b.broadcast_borrow_management()
                else:
                    messages.error(request, 'Quantity cannot be negative.')
            except ValueError:
                messages.error(request, 'Invalid quantity value.')
        else:
            messages.error(request, 'No quantity provided.')
        return redirect('index')

    return redirect('index')


# ─────────────────────────────────────────────────────────────────────────────
#  Borrow requests
# ─────────────────────────────────────────────────────────────────────────────

@login_required
@no_cache
def borrow_requests(request):
    if request.user.role != 'staff':
        raise PermissionDenied
    pending       = BorrowRequest.objects.filter(status='pending').order_by('-created_at')
    pending_count = pending.count()
    return render(request, 'inventory/borrow_requests.html', {
        'pending':       pending,
        'pending_count': pending_count,
    })

@ensure_csrf_cookie
def borrow_item_public(request):
    if request.method == 'POST':
        form = BorrowRequestForm(request.POST)
        if form.is_valid():
            br = form.save(commit=False)
            br.save()
            messages.success(request, 'Your request has been submitted. Staff will review it soon.')
            return redirect('borrow_item_public')
    else:
        form = BorrowRequestForm()
    
    return render(request, 'inventory/borrow_item.html', {
        'form': form,
        'available_items': Item.objects.filter(available_quantity__gt=0),
    })

@login_required
@no_cache
def borrow_management(request):
    if request.user.role != 'staff':
        raise PermissionDenied
    # Table data loads via AJAX after paint to avoid huge HTML.
    return render(request, 'inventory/borrow_management.html', {})


@login_required
@no_cache
def device_monitoring(request):
    if request.user.role != 'staff':
        raise PermissionDenied
    # Rows load via AJAX after paint to avoid huge HTML.
    return render(request, 'inventory/device_monitoring.html', {
        'rows_json': '[]',
    })
 
 
# ─────────────────────────────────────────────────────────────────────────────
#  Graduation Warnings
# ─────────────────────────────────────────────────────────────────────────────

@login_required
@no_cache
def graduation_warnings(request):
    if request.user.role != 'staff':
        raise PermissionDenied

    graduating_keywords = ['4th', 'fourth', '5th', 'fifth']

    from django.db.models import Prefetch
    active_transactions = Transaction.objects.select_related(
        'item', 'borrower', 'borrow_request'
    ).prefetch_related(
        Prefetch('devices', queryset=TransactionDevice.objects.all())
    ).filter(
        status='borrowed',
        borrow_request__borrower_type='student',
    ).order_by('-borrowed_at')

    warnings = []
    for tx in active_transactions:
        br = tx.borrow_request
        if not br:
            continue
        year_level = (br.year_level or '').strip().lower()
        if not year_level:
            year_level = (br.year_section or '').strip().lower()
        if not any(k in year_level for k in graduating_keywords):
            continue

        qty_outstanding = tx.quantity_borrowed - tx.returned_qty
        all_devices = list(tx.devices.all())
        serials_display = ', '.join(d.serial_number for d in all_devices) if all_devices else (tx.serial_number or '—')

        warnings.append({
            'borrower_name':   br.borrower_name,
            'year_level':      br.year_level or br.year_section or '—',
            'section':         br.section or '—',
            'college':         br.college or br.office_college or '—',
            'academic_year':   br.academic_year or '—',
            'student_id':      br.student_id or '—',
            'item_name':       tx.item.name,
            'qty_outstanding': qty_outstanding,
            'serial_number':   serials_display,
            'borrowed_at':     format_ph_time(tx.borrowed_at),
            'officer':         (tx.borrower.get_full_name() or '').strip() or tx.borrower.username,
            'tx_id':           br.transaction_id,
        })

    pending_count = BorrowRequest.objects.filter(status='pending').count()

    return render(request, 'inventory/graduation_warnings.html', {
        'warnings':      warnings,
        'warning_count': len(warnings),
        'pending_count': pending_count,
    })


# ─────────────────────────────────────────────────────────────────────────────
#  Device Monitoring Save / Delete
# ─────────────────────────────────────────────────────────────────────────────

@login_required
@require_POST
def device_monitoring_save(request):
    if request.user.role != 'staff':
        raise PermissionDenied

    if request.content_type == 'application/json':
        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            return JsonResponse({'ok': False, 'error': 'Invalid JSON'}, status=400)

        rows = data.get('rows', [])
        saved_count = 0
        errors = []

        for row_data in rows:
            row_id = row_data.get('row_id')
            if not row_id:
                continue

            fields = {
                'box_number': row_data.get('box_number', ''),
                'office_college': row_data.get('office_college', ''),
                'assigned_mr': row_data.get('assigned_mr', ''),
                'accountable_person': row_data.get('accountable_person', ''),
                'borrower_type': row_data.get('borrower_type', ''),
                'accountable_officer': row_data.get('accountable_officer', ''),
                'device': row_data.get('device', 'Tablet'),
                'serial_number': row_data.get('serial_number', ''),
                'serviceable': row_data.get('serviceable') == 'on',
                'non_serviceable': row_data.get('non_serviceable') == 'on',
                'sealed': row_data.get('sealed') == 'on',
                'missing': row_data.get('missing') == 'on',
                'incomplete': row_data.get('incomplete') == 'on',
                'ptr': row_data.get('ptr', ''),
                'remarks': row_data.get('remarks', ''),
                'issue': row_data.get('issue', ''),
            }

            try:
                if row_id == 'new':
                    DeviceMonitor.objects.create(**fields)
                    saved_count += 1
                else:
                    obj = DeviceMonitor.objects.get(pk=int(row_id))
                    old_date_returned = obj.date_returned
                    for attr, value in fields.items():
                        setattr(obj, attr, value)
                    obj.date_returned = old_date_returned
                    obj.save()
                    saved_count += 1
            except Exception as e:
                errors.append(f"Row {row_id}: {str(e)}")

        b = _broadcasts()
        b.broadcast_device_monitoring()
        b.broadcast_dashboard()

        return JsonResponse({
            'ok': True,
            'saved': saved_count,
            'errors': errors
        })

    # Normal form submission (individual row saves)
    ids                  = request.POST.getlist('row_id')
    box_numbers          = request.POST.getlist('box_number')
    offices              = request.POST.getlist('office_college')
    assigned_mr_list     = request.POST.getlist('assigned_mr')
    accountables         = request.POST.getlist('accountable_person')
    borrower_types       = request.POST.getlist('borrower_type')
    accountable_officers = request.POST.getlist('accountable_officer')
    devices              = request.POST.getlist('device')
    serials              = request.POST.getlist('serial_number')
    serviceables         = request.POST.getlist('serviceable')
    non_serviceables     = request.POST.getlist('non_serviceable')
    sealeds              = request.POST.getlist('sealed')
    missings             = request.POST.getlist('missing')
    ptr_list             = request.POST.getlist('ptr')
    incompletes          = request.POST.getlist('incomplete')
    remarks_list         = request.POST.getlist('remarks')
    issue_list           = request.POST.getlist('issue')

    for i, row_id in enumerate(ids):
        def get(lst, idx=i):
            return lst[idx] if idx < len(lst) else ''

        fields = dict(
            box_number          = get(box_numbers),
            office_college      = get(offices),
            assigned_mr         = get(assigned_mr_list),
            accountable_person  = get(accountables),
            borrower_type       = get(borrower_types),
            accountable_officer = get(accountable_officers),
            device              = get(devices) or 'Tablet',
            serial_number       = get(serials),
            serviceable         = get(serviceables)     == 'on',
            non_serviceable     = get(non_serviceables) == 'on',
            sealed              = get(sealeds)          == 'on',
            missing             = get(missings)         == 'on',
            incomplete          = get(incompletes)      == 'on',
            ptr                 = get(ptr_list),
            remarks             = get(remarks_list),
            issue               = get(issue_list),
        )

        if row_id == 'new':
            DeviceMonitor.objects.create(**fields)
        else:
            try:
                obj = DeviceMonitor.objects.get(pk=int(row_id))
                existing_date_returned = obj.date_returned
                for attr, val in fields.items():
                    setattr(obj, attr, val)
                obj.date_returned = existing_date_returned
                obj.save()
            except DeviceMonitor.DoesNotExist:
                pass

    b = _broadcasts()
    b.broadcast_device_monitoring()
    b.broadcast_dashboard()

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'ok': True, 'saved': len(ids)})

    return redirect('device_monitoring')


@login_required
@require_POST
def device_monitoring_delete(request, row_id):
    if request.user.role != 'staff':
        raise PermissionDenied
    obj = get_object_or_404(DeviceMonitor, pk=row_id)
    obj.delete()
    b = _broadcasts()
    b.broadcast_device_monitoring()
    b.broadcast_dashboard()
    return redirect('device_monitoring')


# ─────────────────────────────────────────────────────────────────────────────
#  Excel Import Helpers & Views
# ─────────────────────────────────────────────────────────────────────────────

def _normalize_header(h):
    h = str(h or '').strip().lower()
    h = re.sub(r'\.', '', h)
    h = re.sub(r'#', '', h)
    h = re.sub(r'\s+', ' ', h).strip()
    return h


@login_required
@require_http_methods(["POST"])
def device_monitoring_import(request):
    if request.user.role != 'staff':
        return JsonResponse({'ok': False, 'error': 'Forbidden'}, status=403)

    excel_file = request.FILES.get('excel_file')
    if not excel_file:
        return JsonResponse({'ok': False, 'error': 'No file uploaded.'}, status=400)
    if not excel_file.name.lower().endswith(('.xlsx', '.xls')):
        return JsonResponse({'ok': False, 'error': 'Invalid file. Please upload .xlsx or .xls.'}, status=400)

    try:
        wb = openpyxl.load_workbook(excel_file, data_only=True)
        ws = wb.active
    except Exception as exc:
        return JsonResponse({'ok': False, 'error': f'Could not read Excel file: {exc}'}, status=400)

    # Detect header row
    header_row_num = None
    all_rows = list(ws.iter_rows(min_row=1, max_row=15, values_only=True))
    for row_idx, row in enumerate(all_rows, start=1):
        non_empty = [c for c in row if c is not None and str(c).strip()]
        if len(non_empty) >= 3:
            header_row_num = row_idx
            break

    if header_row_num is None:
        return JsonResponse({'ok': False, 'error': 'Could not find a header row (need ≥ 3 filled cells).'}, status=400)

    header_row = all_rows[header_row_num - 1]

    ALIASES = {
        'box number':           'box_number',
        'box no':               'box_number',
        'box':                  'box_number',
        'serial number':        'serial_number',
        'serial no':            'serial_number',
        'serial':               'serial_number',
        'sn':                   'serial_number',
        'college office':       'office_college',
        'college':              'office_college',
        'office':               'office_college',
        'name of student':      'accountable_person',
        'student name':         'accountable_person',
        'name':                 'accountable_person',
        'accountable person':   'accountable_person',
        'borrower type':        'borrower_type',
        'type':                 'borrower_type',
        'accountable officer':  'accountable_officer',
        'officer':              'accountable_officer',
        'assigned mr':          'assigned_mr',
        'mr':                   'assigned_mr',
        'assigned m r':         'assigned_mr',
        'device':               'device',
        'ptr':                  'ptr',
        'status':               'release_status_import',
        'release return':       'release_status_import',
        'release  return':      'release_status_import',
        'released return':      'release_status_import',
        'released  return':     'release_status_import',
        'released returned':    'release_status_import',
        'released  returned':   'release_status_import',
        'release status':       'release_status_import',
        'return status':        'release_status_import',
        'date returned':        'date_returned',
        'date released':        'date_returned',
        'remarks':              'remarks',
        'issue':                'issue',
    }

    def _norm(h):
        import re as _re
        h = str(h or '').strip().lower()
        h = _re.sub(r'[.#/\\\-_]', ' ', h)
        h = _re.sub(r'\s+', ' ', h).strip()
        return h

    ALIASES_PRIMARY = {k: v for k, v in ALIASES.items() if k != 'status'}
    col_map = {}
    for col_idx, cell_val in enumerate(header_row):
        norm = _norm(cell_val)
        field = ALIASES_PRIMARY.get(norm)
        if field and field not in col_map.values():
            col_map[col_idx] = field
    if 'release_status_import' not in col_map.values():
        for col_idx, cell_val in enumerate(header_row):
            if _norm(cell_val) == 'status' and col_idx not in col_map:
                col_map[col_idx] = 'release_status_import'
                break
    if 'release_status_import' not in col_map.values():
        for col_idx, cell_val in enumerate(header_row):
            raw = str(cell_val).strip().lower()
            norm = _norm(cell_val)
            if 'release' in raw or 'release' in norm:
                if col_idx not in col_map:
                    col_map[col_idx] = 'release_status_import'
                    break

    if 'serial_number' not in col_map.values():
        return JsonResponse({
            'ok': False,
            'error': f'Could not find a "Serial Number" column. Headers: {[str(h) for h in header_row if h]}'
        }, status=400)

    rows_data = []
    data_rows = list(ws.iter_rows(min_row=header_row_num + 1, values_only=True))
    for row in data_rows:
        if all(c is None or str(c).strip() == '' for c in row):
            continue

        str_data = {}
        raw_data = {}
        for col_idx, field_name in col_map.items():
            raw = row[col_idx] if col_idx < len(row) else None
            raw_data[field_name] = raw
            str_data[field_name] = str(raw).strip() if raw is not None else ''

        serial = str_data.get('serial_number', '').strip()
        if not serial:
            continue

        release_norm = _norm(str_data.get('release_status_import', ''))
        is_returned = (
            release_norm in ('returned', 'return')
            or release_norm.startswith('returned')
        )
        is_released = (
            bool(release_norm)
            and not is_returned
            and (
                release_norm in ('released', 'release', 'borrowed', 'out')
                or release_norm.startswith('released')
                or release_norm.startswith('release return')
            )
        )

        bt = str_data.get('borrower_type', '').strip().lower()
        borrower_type = 'employee' if any(k in bt for k in ('employee', 'emp', 'staff')) else 'student'

        raw_date = raw_data.get('date_returned')
        if raw_date is not None and not isinstance(raw_date, str):
            raw_date = str(raw_date)

        box_number_raw = str_data.get('box_number', '').strip()
        if box_number_raw.endswith('.0') and '.' not in box_number_raw[:-2] and box_number_raw[:-2].isdigit():
            box_number_raw = box_number_raw[:-2]

        rows_data.append({
            'serial_number':       serial,
            'box_number':          box_number_raw,
            'office_college':      str_data.get('office_college', ''),
            'accountable_person':  str_data.get('accountable_person', ''),
            'borrower_type':       borrower_type,
            'accountable_officer': str_data.get('accountable_officer', ''),
            'assigned_mr':         str_data.get('assigned_mr', ''),
            'device':              str_data.get('device', '') or 'Tablet',
            'ptr':                 str_data.get('ptr', ''),
            'remarks':             str_data.get('remarks', ''),
            'issue':               str_data.get('issue', ''),
            'date_returned_raw':   raw_date or '',
            'is_returned':         is_returned,
            'is_released':         is_released,
        })

    if not rows_data:
        return JsonResponse({
            'ok': True, 'task_id': None,
            'total': 0, 'created': 0, 'updated': 0,
            'message': 'No data rows found in the file.',
        })

    try:
        from inventory.tasks import process_excel_import
        task = process_excel_import.delay(rows_data, request.user.id)
        task_id = task.id
    except Exception:
        try:
            from inventory.tasks import process_excel_import
            result = process_excel_import(rows_data, request.user.id)
            return JsonResponse({
                'ok':      True,
                'task_id': None,
                'total':   len(rows_data),
                'created': result.get('created', 0),
                'updated': result.get('updated', 0),
                'errors':  result.get('errors', []),
                'message': 'Import complete (ran synchronously).',
                'done':    True,
            })
        except Exception as exc2:
            return JsonResponse({'ok': False, 'error': str(exc2)}, status=500)

    return JsonResponse({
        'ok':      True,
        'task_id': task_id,
        'total':   len(rows_data),
        'message': f'Import started for {len(rows_data)} row(s).',
    })


@login_required
def import_task_status(request, task_id):
    if request.user.role != 'staff':
        return JsonResponse({'ok': False, 'error': 'Forbidden'}, status=403)

    try:
        from celery.result import AsyncResult
        result = AsyncResult(task_id)
        state  = result.state
    except Exception as exc:
        return JsonResponse({'state': 'FAILURE', 'error': str(exc)})

    if state == 'SUCCESS':
        info = result.result or {}
        if not isinstance(info, dict):
            info = {}
        return JsonResponse({
            'state':   'SUCCESS',
            'ok':      info.get('ok', True),
            'created': info.get('created', 0),
            'updated': info.get('updated', 0),
            'errors':  info.get('errors', []),
        })

    if state == 'FAILURE':
        error_msg = str(result.result) if result.result else 'Unknown error'
        return JsonResponse({'state': 'FAILURE', 'error': error_msg})

    meta = {}
    try:
        if isinstance(result.info, dict):
            meta = result.info
    except Exception:
        pass

    return JsonResponse({
        'state':    state,
        'progress': meta.get('progress', 0),
        'message':  meta.get('message', 'Processing…'),
    })


# ─────────────────────────────────────────────────────────────────────────────
#  Staff Borrow Confirmation / Decline
# ─────────────────────────────────────────────────────────────────────────────

@login_required
@no_cache
def staff_confirm_borrow(request, request_id):
    if request.user.role != 'staff':
        raise PermissionDenied
    borrow_req = get_object_or_404(BorrowRequest, id=request_id, status='pending')

    if request.method == 'POST':
        form = StaffBorrowForm(request.POST)
        if form.is_valid():
            serial_numbers = form.cleaned_data['serial_numbers']
            box_numbers    = form.cleaned_data['box_numbers']
            quantity       = form.cleaned_data['quantity_borrowed']

            transaction = form.save(commit=False)
            transaction.borrower       = request.user
            transaction.borrow_request = borrow_req
            transaction.office_college = borrow_req.office_college
            transaction.status         = 'borrowed'
            transaction.serial_number  = ', '.join(serial_numbers)
            transaction.item.available_quantity -= quantity
            transaction.item.save()
            transaction.save()

            borrow_req.status = 'accepted'
            borrow_req.save()

            accountable_officer = request.user.get_full_name() or request.user.username
            assigned_mr = request.POST.get('assigned_mr', '').strip()

            device_monitors = []
            for i, serial in enumerate(serial_numbers):
                box = box_numbers[i] if i < len(box_numbers) else ''
                TransactionDevice.objects.create(
                    transaction=transaction,
                    serial_number=serial,
                    box_number=box,
                    returned=False,
                    returned_at=None,
                )
                device_monitors.append(DeviceMonitor(
                    box_number=box,
                    office_college=borrow_req.office_college,
                    accountable_person=borrow_req.borrower_name,
                    borrower_type=borrow_req.borrower_type,
                    accountable_officer=accountable_officer,
                    device=transaction.item.name,
                    serial_number=serial,
                    serviceable=True,
                    non_serviceable=False,
                    sealed=False,
                    missing=False,
                    incomplete=False,
                    assigned_mr=assigned_mr,
                ))

            DeviceMonitor.objects.bulk_create(device_monitors)

            b = _broadcasts()
            b.broadcast_all()
            return redirect('index')
    else:
        form = StaffBorrowForm(initial={
            'quantity_borrowed': borrow_req.quantity,
            'office_college':    borrow_req.office_college,
        })

    return render(request, 'inventory/staff_confirm_borrow.html', {
        'form':       form,
        'borrow_req': borrow_req,
    })


@login_required
@no_cache
def decline_request(request, request_id):
    if request.user.role != 'staff':
        raise PermissionDenied
    borrow_req = get_object_or_404(BorrowRequest, id=request_id, status='pending')
    if request.method == 'POST':
        borrow_req.status = 'declined'
        borrow_req.save()
        b = _broadcasts()
        b.broadcast_borrow_requests()
        b.broadcast_dashboard()
    return redirect('borrow_requests')


# ─────────────────────────────────────────────────────────────────────────────
#  Return / Condition
# ─────────────────────────────────────────────────────────────────────────────

@login_required
def return_item(request, transaction_id):
    if request.user.role != 'staff':
        raise PermissionDenied
    transaction = get_object_or_404(Transaction, id=transaction_id)
    if request.method == 'POST' and transaction.status != 'returned':
        transaction.status = 'returned'
        transaction.returned_at = get_ph_time()
        transaction.save()
        b = _broadcasts()
        b.broadcast_borrow_management()
        b.broadcast_dashboard()
        return redirect('borrow_management')
    return render(request, 'inventory/return_item.html', {'transaction': transaction})


@login_required
@no_cache
def update_condition(request, transaction_id):
    if request.user.role != 'staff':
        raise PermissionDenied
    tx = get_object_or_404(Transaction, id=transaction_id)
    if request.method == 'POST':
        form = TransactionConditionForm(request.POST, instance=tx)
        if form.is_valid():
            form.save()
            b = _broadcasts()
            b.broadcast_borrow_management()
    return redirect('borrow_management')


@login_required
@require_POST
def update_returned_qty(request, transaction_id):
    if request.user.role != 'staff':
        return JsonResponse({'error': 'Forbidden'}, status=403)

    tx = get_object_or_404(Transaction, id=transaction_id)

    try:
        new_returned = int(request.POST.get('returned_qty', 0))
    except (ValueError, TypeError):
        return JsonResponse({'error': 'Invalid value'}, status=400)

    new_returned = max(0, min(new_returned, tx.quantity_borrowed))
    delta        = new_returned - tx.returned_qty

    if delta != 0:
        tx.item.available_quantity = max(0, tx.item.available_quantity + delta)
        tx.item.save()

    tx.returned_qty = new_returned
    tx.returned_at = get_ph_time() if new_returned > 0 else None
    tx.status = 'returned' if new_returned >= tx.quantity_borrowed else 'borrowed'
    tx.save()

    b = _broadcasts()
    b.broadcast_borrow_management()
    b.broadcast_dashboard()

    items         = Item.objects.all()
    available_qty = sum(i.available_quantity for i in items)
    agg           = Transaction.objects.annotate(
        still_out=ExpressionWrapper(
            F('quantity_borrowed') - F('returned_qty'),
            output_field=IntegerField()
        )
    ).aggregate(total=Sum('still_out'))
    borrowed_qty = max(0, agg['total'] or 0)

    return JsonResponse({
        'ok':             True,
        'returned_qty':   tx.returned_qty,
        'status':         tx.status,
        'returned_at':    format_ph_time(tx.returned_at),
        'fully_returned': tx.returned_qty >= tx.quantity_borrowed,
        'pie': {'available': available_qty, 'borrowed': borrowed_qty},
    })


@login_required
@require_POST
def return_devices(request, transaction_id):
    if request.user.role != 'staff':
        return JsonResponse({'error': 'Forbidden'}, status=403)

    tx = get_object_or_404(Transaction, id=transaction_id)

    try:
        body = json.loads(request.body)
    except (json.JSONDecodeError, AttributeError):
        return JsonResponse({'error': 'Invalid JSON'}, status=400)

    device_ids = body.get('device_ids', [])
    serials    = body.get('serials', [])

    now_ph = get_ph_time()
    returned_serials = []

    if device_ids:
        real_ids = [d for d in device_ids if d is not None]
        if real_ids:
            updated_devices = TransactionDevice.objects.filter(
                id__in=real_ids,
                transaction=tx,
                returned=False,
            )
            returned_serials = list(updated_devices.values_list('serial_number', flat=True))
            updated_devices.update(returned=True, returned_at=now_ph)
    elif serials:
        for sn in serials:
            td = tx.devices.filter(serial_number=sn, returned=False).first()
            if td:
                td.returned    = True
                td.returned_at = now_ph
                td.save()
                returned_serials.append(sn)

    if returned_serials:
        if tx.borrow_request:
            borrower_name = tx.borrow_request.borrower_name
            office        = tx.borrow_request.office_college
        else:
            borrower_name = tx.borrower.get_full_name() or tx.borrower.username
            office        = tx.office_college

        DeviceMonitor.objects.filter(
            serial_number__in=returned_serials,
            accountable_person=borrower_name,
            office_college=office,
            date_returned__isnull=True,
        ).update(date_returned=now_ph)

    if tx.devices.exists():
        returned_count = tx.devices.filter(returned=True).count()
    else:
        returned_count = tx.returned_qty + len(returned_serials)

    returned_count = min(returned_count, tx.quantity_borrowed)

    delta = returned_count - tx.returned_qty
    if delta > 0:
        tx.item.available_quantity = tx.item.available_quantity + delta
        tx.item.save()

    tx.returned_qty = returned_count
    tx.returned_at  = now_ph if returned_count > 0 else tx.returned_at
    tx.status       = 'returned' if returned_count >= tx.quantity_borrowed else 'borrowed'
    tx.save()

    b = _broadcasts()
    b.broadcast_borrow_management()
    b.broadcast_dashboard()
    b.broadcast_device_monitoring()

    return JsonResponse({
        'ok':            True,
        'returned_qty':  tx.returned_qty,
        'status':        tx.status,
        'fully_returned': tx.returned_qty >= tx.quantity_borrowed,
        'returned_at':   format_ph_time(tx.returned_at),
    })


# ─────────────────────────────────────────────────────────────────────────────
#  Graduation Warnings (again? already defined above, so we skip duplicate)
#  (Already provided once above)
# ─────────────────────────────────────────────────────────────────────────────


# ═══════════════════════════════════════════════════════════════════════════
#  Excel Export Dispatchers  (tasks are in inventory/tasks.py)
# ═══════════════════════════════════════════════════════════════════════════

@login_required
def export_borrow_management(request):
    if request.user.role not in ('staff', 'admin'):
        raise PermissionDenied
    from .tasks import generate_borrow_management_export
    task = generate_borrow_management_export.delay(request.user.id)
    return JsonResponse({'ok': True, 'task_id': task.id})


@login_required
def export_device_monitoring(request):
    if request.user.role not in ('staff', 'admin'):
        raise PermissionDenied
    from .tasks import generate_device_monitoring_export
    task = generate_device_monitoring_export.delay(request.user.id)
    return JsonResponse({'ok': True, 'task_id': task.id})


# ─────────────────────────────────────────────────────────────────────────────
#  Download + export status views  (uses cache set by Celery tasks)
# ─────────────────────────────────────────────────────────────────────────────

@login_required
def download_export(request, token):
    file_data = cache.get(f'export_{token}')
    filename = cache.get(f'export_{token}_fn', 'export.xlsx')
    if not file_data:
        return HttpResponse('Export expired or not found.', status=404)

    resp = HttpResponse(
        file_data,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    resp['Content-Disposition'] = f'attachment; filename="{filename}"'
    return resp


@login_required
def export_task_status(request, task_id):
    from celery.result import AsyncResult
    result = AsyncResult(task_id)
    info = result.result if result.ready() else {}
    token = info.get('token', '') if isinstance(info, dict) else ''
    return JsonResponse({'state': result.state, 'token': token})


# ─────────────────────────────────────────────────────────────────────────────
#  Database keep‑alive (Neon)
# ─────────────────────────────────────────────────────────────────────────────

@csrf_exempt
def db_keepalive(request):
    try:
        from django.db import connection
        with connection.cursor() as cursor:
            cursor.execute('SELECT 1')
        return JsonResponse({'ok': True})
    except Exception as exc:
        return JsonResponse({'ok': False, 'error': str(exc)}, status=500)