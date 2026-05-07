"""
inventory/tasks.py

Celery task for async Excel import of DeviceMonitor records.
Handles create/update per serial number, marks returned/released devices,
and broadcasts live updates when done.

Bulk operations are chunked at CHUNK_SIZE rows to prevent Neon Postgres
from timing out on large imports (4k+ rows).
"""
import random
import pytz
from datetime import datetime, date as _date

from celery import shared_task
from django.utils import timezone
from inventory.consumers import _fmt_ph


PH_TZ      = pytz.timezone('Asia/Manila')
CHUNK_SIZE = 500   # rows per bulk_create / bulk_update call


def _get_ph_time():
    return timezone.now().astimezone(PH_TZ)


def _parse_excel_date(raw):
    """
    Convert an openpyxl cell value to a PH-timezone-aware datetime.
    Returns None for blank / unparseable values.
    """
    if raw is None or str(raw).strip() in ('', '—', '-', 'N/A', 'None'):
        return None
    if isinstance(raw, datetime):
        return raw if raw.tzinfo else PH_TZ.localize(raw)
    if isinstance(raw, _date) and not isinstance(raw, datetime):
        return PH_TZ.localize(datetime(raw.year, raw.month, raw.day))
    text = str(raw).strip()
    for fmt in (
        '%Y-%m-%d %H:%M:%S', '%Y-%m-%d %H:%M', '%Y-%m-%d',
        '%m/%d/%Y %H:%M:%S', '%m/%d/%Y %H:%M', '%m/%d/%Y',
        '%d/%m/%Y %H:%M:%S', '%d/%m/%Y %H:%M', '%d/%m/%Y',
        '%b %d, %Y %I:%M %p', '%b %d, %Y',
        '%B %d, %Y %I:%M %p', '%B %d, %Y',
    ):
        try:
            return PH_TZ.localize(datetime.strptime(text, fmt))
        except ValueError:
            continue
    return None


def _chunked_bulk_create(model, objects, chunk_size=CHUNK_SIZE, ignore_conflicts=False):
    """bulk_create in chunks to avoid Neon query size / timeout limits."""
    created = 0
    for i in range(0, len(objects), chunk_size):
        chunk = objects[i:i + chunk_size]
        model.objects.bulk_create(chunk, ignore_conflicts=ignore_conflicts)
        created += len(chunk)
    return created


def _chunked_bulk_update(model, objects, fields, chunk_size=CHUNK_SIZE):
    """bulk_update in chunks to avoid Neon query size / timeout limits."""
    updated = 0
    for i in range(0, len(objects), chunk_size):
        chunk = objects[i:i + chunk_size]
        model.objects.bulk_update(chunk, fields)
        updated += len(chunk)
    return updated


@shared_task(bind=True, soft_time_limit=300, time_limit=360)
def process_excel_import(self, rows_data, user_id):
    """
    rows_data : list of dicts already parsed from the Excel file by the view.
    user_id   : PK of the staff user who triggered the import.

    Each dict has keys:
        serial_number, box_number, office_college, accountable_person,
        borrower_type, accountable_officer, assigned_mr, device, ptr,
        remarks, issue,
        is_returned (bool), is_released (bool),
        date_returned_raw (raw cell value – passed through _parse_excel_date here)
    """
    from django.contrib.auth import get_user_model
    from inventory.models import (
        Item, DeviceMonitor, TransactionDevice,
        BorrowRequest, Transaction,
    )

    User = get_user_model()
    try:
        user = User.objects.get(pk=user_id)
    except User.DoesNotExist:
        return {'ok': False, 'error': f'User {user_id} not found'}

    # ── Get the staff's full name for accountable_officer column ────────────
    staff_name = user.get_full_name().strip() or user.username

    # ── Wake up Neon free tier before heavy queries ──────────────────────────
    try:
        from django.db import connection
        connection.ensure_connection()
    except Exception:
        pass

    now_ph = _get_ph_time()

    # Ensure a dummy item exists for "released via import" transactions
    dummy_item, _ = Item.objects.get_or_create(
        name='Tablet (Import)',
        defaults={'quantity': 0, 'available_quantity': 0},
    )

    # ── Batch-fetch existing DeviceMonitor rows ──────────────────────────────
    serials = [d['serial_number'] for d in rows_data if d.get('serial_number')]
    existing_map = {
        dm.serial_number: dm
        for dm in DeviceMonitor.objects.filter(serial_number__in=serials)
    }

    to_create        = []
    to_update        = []
    returned_serials = []
    released_rows    = []
    errors           = []

    for d in rows_data:
        serial = (d.get('serial_number') or '').strip()
        if not serial:
            continue

        # FIX: Extract box_number from the row data before using it
        box_number = (d.get('box_number') or '').strip()

        is_returned   = bool(d.get('is_returned'))
        is_released   = bool(d.get('is_released'))
        date_returned = _parse_excel_date(d.get('date_returned_raw'))

        # If flagged returned but no date, use now
        if is_returned and not date_returned:
            date_returned = now_ph
        # If flagged released, clear any date_returned
        if is_released:
            date_returned = None

        # Clean up .0 suffix that sometimes appears when Excel stores numbers as floats
        if box_number.endswith('.0') and box_number[:-2].isdigit():
            box_number = box_number[:-2]

        defaults = {
            'box_number': box_number,
            'office_college':      (d.get('office_college')      or 'Unknown').strip(),
            'accountable_person':  (d.get('accountable_person')  or '').strip(),
            'borrower_type':       (d.get('borrower_type')       or '').strip().lower(),
            'accountable_officer': staff_name,   # ← OVERRIDE: always use importing staff's name
            'assigned_mr':         (d.get('assigned_mr')         or '').strip(),
            'device':              (d.get('device')              or 'Tablet').strip(),
            'ptr':                 (d.get('ptr')                 or '').strip(),
            'remarks':             (d.get('remarks')             or '').strip(),
            'issue':               (d.get('issue')               or '').strip(),
            'date_returned':       date_returned,
            'is_released':         is_released and not is_returned,
            # Checkboxes always cleared on import
            'serviceable':     False,
            'non_serviceable': False,
            'sealed':          False,
            'missing':         False,
            'incomplete':      False,
        }

        obj = existing_map.get(serial)
        if obj is None:
            to_create.append(DeviceMonitor(serial_number=serial, **defaults))
        else:
            for attr, val in defaults.items():
                setattr(obj, attr, val)
            to_update.append(obj)

        if is_returned:
            returned_serials.append(serial)
        elif is_released:
            released_rows.append(d)

    # ── Chunked bulk DB writes ───────────────────────────────────────────────
    update_fields = [
        'box_number', 'office_college', 'accountable_person', 'borrower_type',
        'accountable_officer', 'assigned_mr', 'device', 'ptr',
        'remarks', 'issue', 'date_returned', 'is_released',
        'serviceable', 'non_serviceable', 'sealed', 'missing', 'incomplete',
    ]

    try:
        if to_create:
            _chunked_bulk_create(DeviceMonitor, to_create)
        if to_update:
            _chunked_bulk_update(DeviceMonitor, to_update, update_fields)
    except Exception as exc:
        errors.append(f'Bulk write error: {exc}')

    # ── Mark returned devices in TransactionDevice ───────────────────────────
    # Process in chunks to avoid huge IN clauses
    if returned_serials:
        for i in range(0, len(returned_serials), CHUNK_SIZE):
            chunk = returned_serials[i:i + CHUNK_SIZE]
            TransactionDevice.objects.filter(
                serial_number__in=chunk,
                returned=False,
            ).update(returned=True, returned_at=now_ph)

    # ── Create BorrowRequest + Transaction rows for Released devices ─────────
    if released_rows:
        released_serials_list = [d['serial_number'] for d in released_rows]

        # Mark any existing TransactionDevice rows as returned in chunks
        for i in range(0, len(released_serials_list), CHUNK_SIZE):
            chunk = released_serials_list[i:i + CHUNK_SIZE]
            TransactionDevice.objects.filter(
                serial_number__in=chunk,
                returned=False,
            ).update(returned=True, returned_at=now_ph)

        # Generate unique 5-digit transaction IDs in one DB call
        existing_ids = set(BorrowRequest.objects.values_list('transaction_id', flat=True))
        new_ids = []
        for _ in released_rows:
            while True:
                tx_id = str(random.randint(10000, 99999))
                if tx_id not in existing_ids:
                    existing_ids.add(tx_id)
                    new_ids.append(tx_id)
                    break

        # Chunk BorrowRequest creation
        borrow_reqs = []
        for i in range(0, len(released_rows), CHUNK_SIZE):
            chunk_rows = released_rows[i:i + CHUNK_SIZE]
            chunk_ids  = new_ids[i:i + CHUNK_SIZE]
            created = BorrowRequest.objects.bulk_create([
                BorrowRequest(
                    transaction_id=chunk_ids[j],
                    borrower_name=(chunk_rows[j].get('accountable_person') or '').strip(),
                    borrower_type=(chunk_rows[j].get('borrower_type') or 'student').strip().lower(),
                    office_college=(chunk_rows[j].get('office_college') or 'Unknown').strip(),
                    college=(chunk_rows[j].get('office_college') or 'Unknown').strip(),
                    item=None,
                    quantity=1,
                    status='accepted',
                    student_id='',
                    year_level='',
                    section='',
                    academic_year='',
                )
                for j in range(len(chunk_rows))
            ])
            borrow_reqs.extend(created)

        # Chunk Transaction creation
        txs = []
        for i in range(0, len(released_rows), CHUNK_SIZE):
            chunk_rows = released_rows[i:i + CHUNK_SIZE]
            chunk_brs  = borrow_reqs[i:i + CHUNK_SIZE]
            created = Transaction.objects.bulk_create([
                Transaction(
                    borrow_request=chunk_brs[j],
                    item=dummy_item,
                    borrower=user,
                    office_college=(chunk_rows[j].get('office_college') or 'Unknown').strip(),
                    quantity_borrowed=1,
                    returned_qty=0,
                    status='borrowed',
                    borrowed_at=now_ph,
                    serial_number=chunk_rows[j]['serial_number'],
                )
                for j in range(len(chunk_rows))
            ])
            txs.extend(created)

        # Chunk TransactionDevice creation
        for i in range(0, len(released_rows), CHUNK_SIZE):
            chunk_rows = released_rows[i:i + CHUNK_SIZE]
            chunk_txs  = txs[i:i + CHUNK_SIZE]
            TransactionDevice.objects.bulk_create([
                TransactionDevice(
                    transaction=chunk_txs[j],
                    serial_number=chunk_rows[j]['serial_number'],
                    box_number=(chunk_rows[j].get('box_number') or '').strip(),
                    returned=False,
                )
                for j in range(len(chunk_rows))
            ])

    # ── Live broadcast ───────────────────────────────────────────────────────
    try:
        from inventory.broadcasts import broadcast_device_monitoring, broadcast_dashboard
        broadcast_device_monitoring()
        broadcast_dashboard()
    except Exception:
        pass  # Don't fail the task just because broadcast errored

    return {
        'ok':      True,
        'created': len(to_create),
        'updated': len(to_update),
        'errors':  errors,
    }

@shared_task(bind=True)
def generate_device_monitoring_export(self, user_id):
    import io
    import re
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from django.core.cache import cache
    from inventory.models import DeviceMonitor, TransactionDevice

    rows = list(DeviceMonitor.objects.all())

    def box_number_key(row):
        bn = row.box_number or ''
        match = re.search(r'(\d+)', bn)
        return (int(match.group(1)), bn) if match else (float('inf'), bn)

    rows.sort(key=box_number_key)

    # ── Annotate release_status ──────────────────────────────────────────────
    # FIX (N+1): pre-fetch all active TransactionDevice rows in one query
    all_serials = [r.serial_number for r in rows if r.serial_number]
    active_td_map = {}
    for td in (TransactionDevice.objects
               .filter(serial_number__in=all_serials, returned=False)
               .select_related('transaction__borrow_request', 'transaction__borrower')):
        if td.serial_number not in active_td_map:
            active_td_map[td.serial_number] = td

    for row in rows:
        if row.date_returned:
            row.release_status = 'Returned'
        elif row.release_status:
            pass  # already set by user via badge — keep it
        else:
            td = active_td_map.get(row.serial_number)
            if td and td.transaction:
                tx = td.transaction
                tx_borrower = (tx.borrow_request.borrower_name
                               if tx.borrow_request
                               else tx.borrower.username)
                if tx_borrower == row.accountable_person and tx.office_college == row.office_college:
                    row.release_status = 'Released'
                else:
                    row.release_status = '—'
            else:
                row.release_status = '—'

    # ── Summary stats ────────────────────────────────────────────────────────
    summary_data        = {}
    device_status_summary = {
        'serviceable': 0, 'non_serviceable': 0, 'sealed': 0,
        'missing': 0, 'incomplete': 0, 'released': 0, 'returned': 0,
    }
    device_type_summary = {}
    mr_stats            = {}

    for row in rows:
        college     = row.office_college or 'Unknown'
        assigned_mr = (row.assigned_mr or '').strip() or '—'

        if college not in summary_data:
            summary_data[college] = {
                'total_devices': 0, 'serviceable': 0, 'non_serviceable': 0,
                'sealed': 0, 'missing': 0, 'incomplete': 0,
                'released': 0, 'returned': 0, 'devices_with_issues': 0,
            }
        summary_data[college]['total_devices'] += 1
        for field in ('serviceable', 'non_serviceable', 'sealed', 'missing', 'incomplete'):
            if getattr(row, field):
                summary_data[college][field]   += 1
                device_status_summary[field]   += 1
                if field in ('non_serviceable', 'missing', 'incomplete'):
                    summary_data[college]['devices_with_issues'] += 1

        rs = row.release_status
        if rs == 'Released':
            summary_data[college]['released'] += 1
            device_status_summary['released'] += 1
        elif rs == 'Returned':
            summary_data[college]['returned'] += 1
            device_status_summary['returned'] += 1

        if assigned_mr not in mr_stats:
            mr_stats[assigned_mr] = {
                'total': 0, 'serviceable': 0, 'non_serviceable': 0, 'sealed': 0,
                'missing': 0, 'incomplete': 0, 'released': 0, 'returned': 0,
                'college_details': {},
            }
        stats = mr_stats[assigned_mr]
        stats['total'] += 1
        for field in ('serviceable', 'non_serviceable', 'sealed', 'missing', 'incomplete'):
            if getattr(row, field):
                stats[field] += 1
        if rs == 'Released': stats['released'] += 1
        elif rs == 'Returned': stats['returned'] += 1

        if college not in stats['college_details']:
            stats['college_details'][college] = {
                'total': 0, 'serviceable': 0, 'non_serviceable': 0, 'sealed': 0,
                'missing': 0, 'incomplete': 0, 'released': 0, 'returned': 0,
            }
        col_stats = stats['college_details'][college]
        col_stats['total'] += 1
        for field in ('serviceable', 'non_serviceable', 'sealed', 'missing', 'incomplete'):
            if getattr(row, field):
                col_stats[field] += 1
        if rs == 'Released': col_stats['released'] += 1
        elif rs == 'Returned': col_stats['returned'] += 1

        device_type_summary[row.device or 'Tablet'] = device_type_summary.get(row.device or 'Tablet', 0) + 1

    total_devices     = len(rows)
    total_issues      = (device_status_summary['non_serviceable']
                         + device_status_summary['missing']
                         + device_status_summary['incomplete'])
    health_percentage = ((total_devices - total_issues) / total_devices * 100) if total_devices > 0 else 0
    svc_pct           = (device_status_summary['serviceable'] / total_devices * 100) if total_devices > 0 else 0

    # ── Build Excel ──────────────────────────────────────────────────────────
    wb = Workbook()

    # Sheet 1: Device Details
    ws_details = wb.active
    ws_details.title = 'Device Details'

    headers    = ['Box Number', 'College / Office', 'Student', 'Borrower Type',
                  'Accountable Officer', 'Assigned M.R.', 'Device', 'Serial Number',
                  'Serviceable', 'Non-Serviceable', 'Sealed', 'Missing', 'Incomplete',
                  'Release / Return', 'Date Returned', 'Remarks', 'Issue']
    col_widths = [15, 20, 24, 12, 24, 18, 14, 20, 14, 16, 10, 10, 12, 16, 22, 28, 28]

    ws_details.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    c = ws_details.cell(row=1, column=1, value='Device Monitoring Report')
    c.font = Font(bold=True, size=14, color='000000')
    c.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws_details.row_dimensions[1].height = 30

    ws_details.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    s = ws_details.cell(row=2, column=1)
    ph_now  = _get_ph_time()
    s.value = f'Generated: {ph_now.strftime("%B %d, %Y %I:%M %p")}'
    s.font  = Font(size=9, color='000000')
    s.fill  = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    s.alignment = Alignment(horizontal='center', vertical='center')
    ws_details.row_dimensions[2].height = 16

    fill_hdr = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
    font_hdr = Font(bold=True, color='000000', size=11)
    bdr      = Border(bottom=Side(style='thin', color='CCCCCC'))
    aln      = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for col, heading in enumerate(headers, start=1):
        cell = ws_details.cell(row=3, column=col, value=heading)
        cell.fill = fill_hdr; cell.font = font_hdr
        cell.border = bdr;    cell.alignment = aln
    ws_details.row_dimensions[3].height = 22

    fill_even  = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    fill_odd   = PatternFill(start_color='F9F9F9', end_color='F9F9F9', fill_type='solid')
    font_base  = Font(color='000000', size=10)
    font_green = Font(color='00e5a0', bold=True, size=10)
    border_row = Border(bottom=Side(style='thin', color='EEEEEE'))
    align_row  = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for i, row in enumerate(rows, start=1):
        borrower_type_display = (
            'Student'  if row.borrower_type == 'student'  else
            'Employee' if row.borrower_type == 'employee' else '—'
        )
        date_ret   = _fmt_ph(row.date_returned) if row.date_returned else '—'
        fill_row   = fill_even if i % 2 == 0 else fill_odd
        bool_flags = {9: row.serviceable, 10: row.non_serviceable, 11: row.sealed,
                      12: row.missing, 13: row.incomplete}
        values = [
            row.box_number or '—', row.office_college or '—',
            row.accountable_person or '—', borrower_type_display,
            row.accountable_officer or '—', row.assigned_mr or '—',
            row.device or 'Tablet', row.serial_number or '—',
            '✓' if row.serviceable     else '—',
            '✓' if row.non_serviceable else '—',
            '✓' if row.sealed          else '—',
            '✓' if row.missing         else '—',
            '✓' if row.incomplete      else '—',
            row.release_status or '—', date_ret,
            row.remarks or '—', row.issue or '—',
        ]
        for col, val in enumerate(values, start=1):
            cell = ws_details.cell(row=i + 3, column=col, value=val)
            cell.fill      = fill_row
            cell.font      = font_green if bool_flags.get(col) else font_base
            cell.border    = border_row
            cell.alignment = align_row

    for col, width in enumerate(col_widths, start=1):
        ws_details.column_dimensions[get_column_letter(col)].width = width
    ws_details.freeze_panes = 'A4'

    # Sheet 2: Summary Report
    ws_summary = wb.create_sheet('Summary Report')

    def write_table(ws, start_row, title, tbl_headers, data_rows, tbl_col_widths=None):
        ws.merge_cells(start_row=start_row, start_column=1,
                       end_row=start_row, end_column=len(tbl_headers))
        tc = ws.cell(row=start_row, column=1, value=title)
        tc.font = Font(bold=True, size=12, color='000000')
        tc.fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
        tc.alignment = Alignment(horizontal='center')
        ws.row_dimensions[start_row].height = 25
        hdr_row = start_row + 1
        fill_h2 = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
        font_h2 = Font(bold=True, color='000000', size=11)
        for col, hdr in enumerate(tbl_headers, start=1):
            cell = ws.cell(row=hdr_row, column=col, value=hdr)
            cell.fill = fill_h2; cell.font = font_h2
            cell.alignment = Alignment(horizontal='center')
            cell.border = Border(bottom=Side(style='thin', color='888888'))
        ws.row_dimensions[hdr_row].height = 20
        fe = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        fo = PatternFill(start_color='F9F9F9', end_color='F9F9F9', fill_type='solid')
        fr = Font(color='000000', size=10)
        for j, row_vals in enumerate(data_rows, start=1):
            fill_r = fe if j % 2 == 0 else fo
            for col, val in enumerate(row_vals, start=1):
                cell = ws.cell(row=hdr_row + j, column=col, value=val)
                cell.fill = fill_r; cell.font = fr
                cell.alignment = Alignment(horizontal='center', wrap_text=True)
                cell.border = Border(bottom=Side(style='thin', color='EEEEEE'))
        if tbl_col_widths:
            for col, width in enumerate(tbl_col_widths, start=1):
                ws.column_dimensions[get_column_letter(col)].width = width
        return hdr_row + len(data_rows) + 1

    current_row = 1

    overall_data = [
        ['Total Devices', total_devices],
        ['Serviceable', f"{device_status_summary['serviceable']} ({svc_pct:.1f}%)"],
        ['Sealed', device_status_summary['sealed']],
        ['Non-Serviceable', device_status_summary['non_serviceable']],
        ['Missing', device_status_summary['missing']],
        ['Incomplete', device_status_summary['incomplete']],
        ['Devices with Issues', total_issues],
        ['Overall Device Health', f"{health_percentage:.1f}%"],
    ]
    current_row = write_table(ws_summary, current_row, '📊 OVERALL INVENTORY STATUS',
                              ['Metric', 'Value'], overall_data, [30, 20])
    current_row += 1

    if len(device_type_summary) > 1:
        current_row = write_table(ws_summary, current_row, '📱 DEVICE TYPE DISTRIBUTION',
                                  ['Device Type', 'Count'],
                                  [[k, v] for k, v in device_type_summary.items()],
                                  [25, 15])
        current_row += 1

    # FIX: initialize detail_headers before the if block so merge_cols never
    # references an undefined variable when detail_data is empty.
    detail_headers = ['Assigned M.R.', 'College / Office', 'Total Devices', 'Serviceable',
                      'Non‑Svc', 'Sealed', 'Missing', 'Incomplete', 'Borrowed', 'Returned', 'Healthy %']
    detail_widths  = [25, 30, 12, 12, 12, 10, 10, 12, 12, 12, 12]

    detail_data = []
    for mr_name in sorted(mr_stats.keys(), key=lambda x: (x == '—', x)):
        for college, col_stats in sorted(mr_stats[mr_name]['college_details'].items()):
            total_in = col_stats['total']
            issues   = col_stats['non_serviceable'] + col_stats['missing'] + col_stats['incomplete']
            hp       = ((total_in - issues) / total_in * 100) if total_in > 0 else 0
            detail_data.append([
                mr_name, college, total_in,
                col_stats['serviceable'], col_stats['non_serviceable'], col_stats['sealed'],
                col_stats['missing'], col_stats['incomplete'],
                col_stats['released'], col_stats['returned'], f"{hp:.1f}%",
            ])
    if detail_data:
        current_row = write_table(ws_summary, current_row,
                                  '🔍 DETAILED BREAKDOWN BY ASSIGNED M.R. AND COLLEGE',
                                  detail_headers, detail_data, detail_widths)
        current_row += 1

    merge_cols = len(detail_headers)  # always defined now

    insights_lines = [f"• Overall device health: {health_percentage:.1f}%",
                      f"• Serviceable rate: {svc_pct:.1f}%"]
    if device_status_summary['missing']       > 0: insights_lines.append(f"⚠️ ALERT: {device_status_summary['missing']} device(s) marked MISSING")
    if device_status_summary['non_serviceable'] > 0: insights_lines.append(f"🔧 {device_status_summary['non_serviceable']} device(s) need repair")
    if device_status_summary['incomplete']    > 0: insights_lines.append(f"📦 {device_status_summary['incomplete']} device(s) are incomplete")
    colleges_issues = [c for c, d in summary_data.items() if d['devices_with_issues'] > 0]
    if colleges_issues: insights_lines.append(f"⚠️ Colleges needing attention: {', '.join(colleges_issues)}")
    if device_status_summary['released']      > 0: insights_lines.append(f"🔄 {device_status_summary['released']} device(s) currently borrowed")

    ws_summary.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=merge_cols)
    cell = ws_summary.cell(row=current_row, column=1, value='💡 KEY INSIGHTS\n' + '\n'.join(insights_lines))
    cell.font = Font(bold=True, size=11, color='000000')
    cell.fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
    cell.alignment = Alignment(wrap_text=True, horizontal='left')
    ws_summary.row_dimensions[current_row].height = 30 + 15 * len(insights_lines)
    current_row += 2

    recs = []
    if device_status_summary['missing']         > 0: recs.append(f"🔴 Conduct physical inventory for {device_status_summary['missing']} missing device(s)")
    if device_status_summary['non_serviceable'] > 0: recs.append(f"🔧 Schedule repair for {device_status_summary['non_serviceable']} non‑serviceable devices")
    if device_status_summary['incomplete']      > 0: recs.append(f"📋 Audit {device_status_summary['incomplete']} incomplete devices")
    for college in colleges_issues:
        recs.append(f"📞 Follow up with {college} ({summary_data[college]['devices_with_issues']} device(s) with issues)")
    if not recs:
        recs.append("✅ All devices in good condition. Continue regular monitoring.")

    ws_summary.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=merge_cols)
    cell = ws_summary.cell(row=current_row, column=1, value='🎯 RECOMMENDATIONS\n' + '\n'.join(recs))
    cell.font = Font(bold=True, size=11, color='000000')
    cell.fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
    cell.alignment = Alignment(wrap_text=True, horizontal='left')
    ws_summary.row_dimensions[current_row].height = 30 + 20 * len(recs)

    # ── Save to cache ─────────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    token = self.request.id
    cache.set(f'export_{token}', buf.getvalue(), 300)
    cache.set(f'export_{token}_fn', 'device_monitoring.xlsx', 300)

    return {'ok': True, 'token': token}